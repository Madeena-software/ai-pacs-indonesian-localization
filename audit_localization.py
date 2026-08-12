from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl

DEFAULT_BASE_URL = "http://124.225.183.175:8361"
DEFAULT_DOCTOR_HASH = "#/doctor"
DEFAULT_LIST_HASH = "#/userTable/%E8%83%B8%E9%83%A8DR"

USERNAME_SELECTORS = (
    'input[name="username"]',
    'input[name*="user" i]',
    'input[placeholder*="username" i]',
    'input[placeholder*="user name" i]',
    'input[placeholder*="Nama Pengguna" i]',
    'input[placeholder*="用户名"]',
    'input[placeholder*="账号"]',
    'input[placeholder*="帐号"]',
    'input[type="text"]',
)
PASSWORD_SELECTORS = (
    'input[type="password"]',
    'input[name="password"]',
    'input[name*="pass" i]',
    'input[placeholder*="Kata Sandi" i]',
)
LOGIN_BUTTON_TEXTS = ("Login", "Sign in", "Log in", "Masuk", "登录", "登入")

TECHNICAL_TERMS = {
    "AI", "PACS", "DICOM", "DR", "CR", "CT", "MRI", "JPEG", "PNG", "PDF",
    "ID", "Madeena", "UMI", "HTML", "URL", "JSON", "HTTP", "HTTPS", "GB", "MB", "KB",
    "B-MODE", "2D", "3D", "4D",
    # Product names — must not be translated
    "Insight", "Insight ChestDR", "Insight QCDR", "Insight Chest DR", "YiZhun", "YiZhun AI-PACS",
    # DICOM & Medical Imaging standard technical terms — preserve in English
    "Window Leveling", "Window Level", "Window Width", "Windowing",
}

KNOWN_TRANSLATIONS = {
    # Chinese UI strings
    "用户名": "Nama Pengguna",
    "密码": "Kata Sandi",
    "登录": "Masuk",
    "登入": "Masuk",
    "胸部DR": "DR Dada",
    "医生": "Dokter",
    "影像加载中…": "Memuat gambar…",

    # English & Mixed UI strings
    "Login": "Masuk",
    "Log in": "Masuk",
    "Sign in": "Masuk",
    "Username": "Nama Pengguna",
    "Password": "Kata Sandi",
    "Doctor": "Dokter",
    "Patient": "Pasien",
    "Status": "Status",
    "Search": "Cari",
    "Reset": "Atur Ulang",
    "Export": "Ekspor",
    "Action": "Tindakan",
    "Detail": "Rincian",
    "Operation": "Operasi",
    "Language": "Bahasa",
    "Download Report": "Unduh Laporan",
    "Generate Report": "Buat Laporan",
    "AI Report": "Laporan AI",
    "Image Report": "Laporan Gambar",
    "Unduh Report": "Unduh Laporan",
    "Sex": "Jenis Kelamin",
    "Age": "Usia",
    "Name": "Nama",
    "Date": "Tanggal",
    "Department": "Departemen",
    "Anonymous Patient": "Pasien Anonim",
    "Next Page": "Halaman Berikutnya",
    "Previous Page": "Halaman Sebelumnya",
    "20 / page": "20 / halaman",
    "Go to": "Menuju",
    "Page Size": "Ukuran Halaman",
    "Full Screen": "Layar Penuh",
    "Favorite": "Favorit",
    "Feedback": "Umpan Balik",
    "Home - AI-PACS": "Beranda - AI-PACS",
    "Indonesian": "Bahasa Indonesia",
    "Disclaimer": "Pernyataan Batasan Tanggung Jawab",

    # Medical & Viewer UI Terms
    "Free Layout": "Tata Letak Bebas",
    "Window Leveling": "Window Leveling",
    "Move": "Pindah",
    "Zoom": "Perbesar",
    "Invert": "Inversi Warna",
    "Rotate": "Putar",
    "Magnifier": "Kaca Pembesar",
    "Spotlight": "Sorot",
    "Length": "Panjang",
    "Lesion Contouring": "Kontur Lesi",
    "More": "Lainnya",
    "Lesion List": "Daftar Lesi",
    "Add": "Tambah",
    "LUNG": "PARU",
    "MEDIASTINUM": "MEDIASTINUM",
    "PLEURA": "PLEURA",
    "RIB": "TULANG RUSUK",
    "Cardiac Shadow": "Bayangan Jantung",
    "Abnormal": "Abnormal",
    "Cardiothoracic Ratio": "Rasio Kardiotoraks",
    "Cardiothoracic ratio is 0.54, please correlate clinically.": "Rasio kardiotoraks adalah 0.54, harap korelasikan secara klinis.",
    "Cardiothoracic ratio is 0.54.": "Rasio kardiotoraks adalah 0.54.",
    "Chest DR Intelligent Analysis": "Analisis Intelijen Chest DR",
    "Insight Chest DR": "Insight Chest DR (Nama Produk)",
    "Imaging Report": "Laporan Pemeriksaan",
    "Imaging Findings": "Temuan Pemeriksaan",
    "Imaging Opinion": "Opini Pemeriksaan",
    "Corner Information": "Informasi Sudut",
    "Lesion Marking": "Penandaan Lesi",
    "Scale": "Skala",
    "Show Foreign Objects": "Tampilkan Benda Asing",
    "Original Image": "Gambar Asli",
    "Submit Results": "Kirim Hasil",
    "Copy": "Salin",
    "Previous": "Sebelumnya",
    "Next": "Selanjutnya",
    "Start date": "Tanggal Mulai",
    "End date": "Tanggal Selesai",
    "All": "Semua",
    "Filter": "Filter",
    "Madeena AI-Assisted Diagnosis System": "Sistem Diagnosis Terbantu AI Madeena",
    "Madeena Intelligent": "Madeena Intelligent",
    "PT. Madeena Karya Indonesia": "PT. Madeena Karya Indonesia",

    # Already Indonesian / Medical terms
    "0 item dipilih": "0 item dipilih",
    "Berhasil": "Berhasil",
    "Efusi Pleura": "Efusi Pleura",
    "Ekspor": "Ekspor",
    "Gagal": "Gagal",
    "Kelainan Jantung": "Kelainan Jantung",
    "Lanjutan": "Lanjutan",
    "Massa/Nodul": "Massa/Nodul",
    "Negatif": "Negatif",
    "Pneumonia": "Pneumonia",
    "Pneumotoraks": "Pneumotoraks",
    "Positif": "Positif",
    "Positif atau Negatif": "Positif atau Negatif",
    "Silakan Masukkan ID/Nama Pasien dan Tekan Enter untuk Mencari": "Silakan Masukkan ID/Nama Pasien dan Tekan Enter untuk Mencari",
    "Tekan Enter setelah memasukkan ID untuk memulai pencarian.": "Tekan Enter setelah memasukkan ID untuk memulai pencarian.",
    "Tidak ada kasus ditemukan": "Tidak ada kasus ditemukan",
    "Total 46 data": "Total 46 data",
    "WW：65536 WL：32768": "WW: 65536 WL: 32768",
    "Waktu penerimaan": "Waktu penerimaan",
    "DR Dada - Platform YiZhun AI-PACS": "DR Dada - Platform YiZhun AI-PACS",

    # UI Icons & Punctuation
    "calendar": "Ikon Kalender",
    "caret-down": "Panah Bawah",
    "caret-up": "Panah Atas",
    "down": "Bawah",
    "ellipsis": "Tanda Hubung (Lainnya)",
    "left": "Kiri",
    "right": "Kanan",
    "swap-right": "Panah Kanan",
    "to": "Sampai",
}


@dataclass
class Finding:
    finding_id: str
    route: str
    module_name: str
    page_url: str
    page_title: str
    element_selector: str
    text_observed: str
    classification: str
    expected_indonesian: str
    quality_note: str
    screenshot_path: str
    fullpage_screenshot_path: str = ""


def find_system_chromium() -> str | None:
    override = os.environ.get("PACS_CHROMIUM_EXECUTABLE", "").strip()
    if override and Path(override).is_file():
        return override

    for name in ("chromium", "chromium-browser", "google-chrome", "google-chrome-stable", "msedge", "microsoft-edge"):
        found = shutil.which(name)
        if found:
            return found

    pw_cache = Path.home() / ".cache" / "ms-playwright"
    if pw_cache.is_dir():
        for candidate in sorted(pw_cache.glob("chromium-*/chrome-linux64/chrome"), reverse=True):
            if candidate.is_file():
                return str(candidate)
        for candidate in sorted(pw_cache.glob("chromium-*/chrome-linux/chrome"), reverse=True):
            if candidate.is_file():
                return str(candidate)

    if os.name == "nt":
        roots = [
            os.environ.get("PROGRAMFILES"),
            os.environ.get("PROGRAMFILES(X86)"),
            os.environ.get("LOCALAPPDATA"),
        ]
        relatives = (
            Path("Google/Chrome/Application/chrome.exe"),
            Path("Microsoft/Edge/Application/msedge.exe"),
        )
        for root in roots:
            if not root:
                continue
            for relative in relatives:
                candidate = Path(root) / relative
                if candidate.is_file():
                    return str(candidate)
    return None


def load_credentials(path: str | Path) -> tuple[str, str]:
    values: dict[str, str] = {}
    credential_path = Path(path)
    if not credential_path.is_file() and credential_path.name == "credential.txt":
        dotenv_path = credential_path.with_name(".env")
        if dotenv_path.is_file():
            credential_path = dotenv_path

    if not credential_path.is_file():
        raise ValueError(f"Credential file not found: {path}")

    for raw_line in credential_path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")

    missing = [
        key
        for key in ("AI_PACS_USERNAME", "AI_PACS_PASSWORD")
        if not values.get(key)
    ]
    if missing:
        raise ValueError(f"Missing credential key(s): {', '.join(missing)}")
    return values["AI_PACS_USERNAME"], values["AI_PACS_PASSWORD"]


def probe_http_origin(base_url: str, timeout: float = 5.0) -> dict[str, Any]:
    url = base_url.rstrip("/") + "/"
    request = urllib.request.Request(
        url,
        method="GET",
        headers={"User-Agent": "AI-PACS-Localization-Probe/1"},
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return {"ok": True, "url": url, "status": getattr(response, "status", None), "error": None}
    except urllib.error.HTTPError as exc:
        return {"ok": True, "url": url, "status": exc.code, "error": None}
    except Exception as exc:
        return {"ok": False, "url": url, "status": None, "error": str(exc)}


def generate_finding_id(route: str, element_selector: str, text_observed: str) -> str:
    key = f"{route}\x00{element_selector}\x00{text_observed}"
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:16]


INDONESIAN_KEYWORDS = {
    "produk", "ini", "hanya", "untuk", "keperluan", "penelitian", "bukan", "aplikasi", "klinis",
    "impor", "analisis", "dan", "penilaian", "citra", "dada", "multi-penyakit", "manajemen",
    "antrian", "pengarsipan", "komputasi", "pasien", "nama", "stasiun", "pengguna", "kata",
    "sandi", "masuk", "keluar", "batal", "simpan", "cari", "atur", "ulang", "tindakan",
    "rincian", "laporan", "gambar", "hasil", "diagnosis", "pengaturan", "sistem", "dokter",
    "jenis", "kelamin", "usia", "tanggal", "departemen", "tabel", "daftar", "id", "pemeriksaan",
    "waktu", "penerimaan", "ditemukan", "kasus", "pilih", "pilihan", "tampilkan", "sembunyikan",
    "informasi", "detail", "total", "data", "halaman", "menu", "kembali", "buka", "tutup", "bantu",
    "bantuan", "lihat", "ubah", "hapus", "edit", "beranda", "item", "dipilih", "berhasil", "gagal"
}


def classify_string(text: str) -> tuple[str, str, str]:
    text_clean = text.strip()
    if not text_clean or len(text_clean) < 2:
        return "technical-term", "", ""

    # Pure numeric or punctuation
    if re.match(r"^[\d\.\,\:\-\/\s\%\#\$\(\)\+\=\>\<]+$", text_clean):
        return "technical-term", "", ""

    # Icon names or tech identifiers
    if text_clean in {
        "eye-invisible", "eye", "lock", "user", "search", "loading", "anticon",
        "calendar", "caret-down", "caret-up", "down", "left", "right", "swap-right", "ellipsis"
    }:
        return "technical-term", "", ""

    # Date string
    if re.match(r"^\d{4}[-/]\d{2}[-/]\d{2}", text_clean) or re.match(r"^\d{2}[-/]\d{2}[-/]\d{4}", text_clean) or re.match(r"^\d{1,2}/\d{1,2}/\d{4}", text_clean):
        return "technical-term", "", ""

    # Patient IDs / Study Codes (e.g., 20-LST-20B_Thorax_PA, 20-LST-20B, etc.)
    if re.search(r"^\d{2}-[A-Z0-9_\-]+", text_clean) or re.search(r"^\d{2}-[A-Z]+", text_clean):
        return "technical-term", "", ""

    # DICOM Technique overlays & Measurements (e.g., NaNkVp, mAs, dGy, WW: WL:, mm, cm)
    if re.search(r"\b(NaNkVp|dGy|mAs)\b|WW:\s*\d+|WL:\s*\d+|\d+mm|\d+cm", text_clean):
        return "technical-term", "", ""

    # Technical terms or upper-case codes
    if text_clean in TECHNICAL_TERMS or re.match(r"^[A-Z0-9_\-]+$", text_clean):
        return "technical-term", "", ""

    # Check for quality issue typos
    if text_clean == "Masukan kata sandi":
        return "quality-issue", "Masukkan kata sandi", "Imperative verb requires 'Masukkan' instead of 'Masukan'."

    def get_expected_and_note(raw_text: str, default_cls: str) -> tuple[str, str]:
        exp = KNOWN_TRANSLATIONS.get(raw_text, raw_text)
        if default_cls == "not-indonesian":
            if bool(re.search(r"[\u4e00-\u9fff]", raw_text)):
                note = "Text is in Chinese and requires translation into Indonesian."
            else:
                note = "Text is in English and requires translation into Indonesian."
        elif default_cls == "mixed":
            note = "Text uses a mix of Indonesian and foreign/English terms."
        elif default_cls == "quality-issue":
            note = "Imperative verb requires 'Masukkan' instead of 'Masukan'."
        else: # uncertain
            note = "Text or UI element requires contextual verification by the localization team."
        return exp, note

    # Check for Chinese characters
    has_chinese = bool(re.search(r"[\u4e00-\u9fff]", text_clean))
    if has_chinese:
        exp, note = get_expected_and_note(text_clean, "not-indonesian")
        return "not-indonesian", exp, note

    # Check for mixed
    if "Unduh Report" in text_clean or "Laporan AI (AI Report)" in text_clean:
        exp, note = get_expected_and_note(text_clean, "mixed")
        return "mixed", exp, note

    # Check if text is in KNOWN_TRANSLATIONS or is already an expected Indonesian translation value
    if text_clean in KNOWN_TRANSLATIONS:
        expected = KNOWN_TRANSLATIONS[text_clean]
        if expected == text_clean:
            return "technical-term", "", ""
        exp, note = get_expected_and_note(text_clean, "not-indonesian")
        return "not-indonesian", exp, note

    if text_clean in KNOWN_TRANSLATIONS.values():
        return "technical-term", "", ""

    # Broad English UI words vocabulary
    english_ui_vocab = {
        "login", "sign", "log", "username", "password", "download", "generate", "report",
        "image", "doctor", "patient", "search", "reset", "export", "operation", "action",
        "detail", "sex", "age", "date", "department", "language", "insight", "user", "system",
        "layout", "window", "leveling", "move", "zoom", "invert", "rotate", "magnifier",
        "spotlight", "length", "lesion", "contouring", "list", "lung", "mediastinum", "pleura",
        "rib", "cardiac", "shadow", "abnormal", "cardiothoracic", "ratio", "imaging", "findings",
        "opinion", "corner", "information", "marking", "scale", "foreign", "objects", "original",
        "submit", "results", "copy", "previous", "next", "start", "end", "all", "filter", "select", "view"
    }

    words = [w.lower().strip(".,()!*#") for w in text_clean.split()]
    if words:
        indonesian_word_count = sum(1 for w in words if w in INDONESIAN_KEYWORDS)
        if (indonesian_word_count / len(words)) >= 0.3:
            return "technical-term", "", ""

        english_match_count = sum(1 for w in words if w in english_ui_vocab)
        if (english_match_count / len(words)) >= 0.3:
            exp, note = get_expected_and_note(text_clean, "not-indonesian")
            return "not-indonesian", exp, note

    exp, note = get_expected_and_note(text_clean, "uncertain")
    return "uncertain", exp, note


def export_to_excel(findings: list[Finding], summary: dict[str, Any], output_path: str | Path) -> None:
    from openpyxl.styles import Alignment

    wb = openpyxl.Workbook()
    out_file = Path(output_path)
    base_dir = out_file.parent

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 1: Findings
    ws_findings = wb.active
    ws_findings.title = "Findings"
    
    findings_headers = [
        "finding_id",
        "route",
        "module_name",
        "page_url",
        "page_title",
        "element_selector",
        "text_observed",
        "classification",
        "expected_indonesian",
        "quality_note",
        "screenshot_image",
        "fullpage_screenshot_image",
    ]
    ws_findings.append(findings_headers)

    ws_findings.column_dimensions["A"].width = 18  # finding_id
    ws_findings.column_dimensions["B"].width = 15  # route
    ws_findings.column_dimensions["C"].width = 22  # module_name
    ws_findings.column_dimensions["D"].width = 45  # page_url
    ws_findings.column_dimensions["E"].width = 30  # page_title
    ws_findings.column_dimensions["F"].width = 25  # element_selector
    ws_findings.column_dimensions["G"].width = 35  # text_observed
    ws_findings.column_dimensions["H"].width = 16  # classification
    ws_findings.column_dimensions["I"].width = 25  # expected_indonesian
    ws_findings.column_dimensions["J"].width = 35  # quality_note
    ws_findings.column_dimensions["K"].width = 22  # screenshot_image
    ws_findings.column_dimensions["L"].width = 22  # fullpage_screenshot_image
    
    for row_idx, f in enumerate(findings, start=2):
        rel_path = f.screenshot_path
        full_rel_path = f.fullpage_screenshot_path

        ws_findings.append([
            f.finding_id,
            f.route,
            f.module_name,
            f.page_url,
            f.page_title,
            f.element_selector,
            f.text_observed,
            f.classification,
            f.expected_indonesian,
            f.quality_note,
            "", # element image placeholder column K
            "", # fullpage image placeholder column L
        ])

        from openpyxl.drawing.image import Image as OpenPyXLImage

        # Embed element thumbnail image if screenshot file exists
        if rel_path:
            full_img_path = base_dir / rel_path
            if full_img_path.is_file():
                try:
                    img = OpenPyXLImage(str(full_img_path))
                    max_size = 120
                    if img.width > max_size or img.height > max_size:
                        scale = min(max_size / img.width, max_size / img.height)
                        img.width = int(img.width * scale)
                        img.height = int(img.height * scale)

                    img.anchor = f"K{row_idx}"
                    ws_findings.add_image(img)
                    ws_findings.row_dimensions[row_idx].height = max(50, int(img.height * 0.75))
                except Exception:
                    pass

        # Embed fullpage thumbnail image if fullpage screenshot file exists
        if full_rel_path:
            full_page_img_path = base_dir / full_rel_path
            if full_page_img_path.is_file():
                try:
                    fp_img = OpenPyXLImage(str(full_page_img_path))
                    fp_max_size = 120
                    if fp_img.width > fp_max_size or fp_img.height > fp_max_size:
                        fp_scale = min(fp_max_size / fp_img.width, fp_max_size / fp_img.height)
                        fp_img.width = int(fp_img.width * fp_scale)
                        fp_img.height = int(fp_img.height * fp_scale)

                    fp_img.anchor = f"L{row_idx}"
                    ws_findings.add_image(fp_img)
                    ws_findings.row_dimensions[row_idx].height = max(
                        ws_findings.row_dimensions[row_idx].height or 50,
                        int(fp_img.height * 0.75)
                    )
                except Exception:
                    pass

    # Apply wrap_text to all cells in Findings sheet
    for row in ws_findings.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment

    # Sheet 2: Summary
    ws_summary = wb.create_sheet(title="Summary")
    summary_headers = [
        "run_timestamp",
        "base_url",
        "routes_visited",
        "strings_inspected",
        "findings_total",
        "findings_not_indonesian",
        "findings_mixed",
        "findings_quality_issue",
        "findings_uncertain",
        "viewer_modal_visited",
    ]
    ws_summary.append(summary_headers)
    ws_summary.append([
        summary.get("run_timestamp", ""),
        summary.get("base_url", ""),
        summary.get("routes_visited", 0),
        summary.get("strings_inspected", 0),
        summary.get("findings_total", 0),
        summary.get("findings_not_indonesian", 0),
        summary.get("findings_mixed", 0),
        summary.get("findings_quality_issue", 0),
        summary.get("findings_uncertain", 0),
        summary.get("viewer_modal_visited", False),
    ])

    # Apply wrap_text to all cells in Summary sheet
    for row in ws_summary.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


class LocalizationAuditor:
    def __init__(
        self,
        base_url: str,
        username: str,
        password: str,
        storage_state: str | Path,
        screenshots_dir: str | Path,
        headed: bool = False,
        timeout_ms: int = 30000,
    ) -> None:
        self.base_url = base_url.rstrip("/")
        self.username = username
        self.password = password
        self.storage_state = Path(storage_state)
        self.screenshots_dir = Path(screenshots_dir)
        self.headed = headed
        self.timeout_ms = timeout_ms

        self.findings: list[Finding] = []
        self.seen_keys: set[tuple[str, str, str]] = set()
        self.strings_inspected = 0
        self.routes_visited: set[str] = set()

        self._playwright = None
        self.browser = None
        self.context = None
        self.page = None

    def start(self) -> None:
        try:
            from playwright.sync_api import sync_playwright
        except ImportError as exc:
            raise RuntimeError(
                "Playwright is not installed. Run: python -m pip install playwright"
            ) from exc

        self._playwright = sync_playwright().start()
        managed_executable = Path(self._playwright.chromium.executable_path)
        launch_kwargs: dict[str, Any] = {"headless": not self.headed}
        if not managed_executable.is_file():
            system_chromium = find_system_chromium()
            if system_chromium:
                launch_kwargs["executable_path"] = system_chromium

        self.browser = self._playwright.chromium.launch(**launch_kwargs)

        context_kwargs: dict[str, Any] = {
            "accept_downloads": True,
            "locale": "id-ID",
            "extra_http_headers": {"Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7"},
        }
        if self.storage_state.is_file():
            try:
                json.loads(self.storage_state.read_text(encoding="utf-8"))
                context_kwargs["storage_state"] = str(self.storage_state)
            except (OSError, json.JSONDecodeError):
                pass

        self.context = self.browser.new_context(**context_kwargs)
        self.context.add_init_script(
            "try { "
            "localStorage.setItem('umi_locale', 'id-ID'); "
            "localStorage.setItem('i18nextLng', 'id'); "
            "localStorage.setItem('locale', 'id-ID'); "
            "localStorage.setItem('lang', 'id'); "
            "} catch(e) {}"
        )
        self.page = self.context.new_page()
        self.page.set_default_timeout(self.timeout_ms)
        self.page.set_default_navigation_timeout(self.timeout_ms)

    def close(self) -> None:
        if self.context:
            try:
                self.context.close()
            except Exception:
                pass
            self.context = None
        if self.browser:
            try:
                self.browser.close()
            except Exception:
                pass
            self.browser = None
        if self._playwright:
            try:
                self._playwright.stop()
            except Exception:
                pass
            self._playwright = None

    def _first_visible_css(self, selectors: Iterable[str]):
        for selector in selectors:
            locator = self.page.locator(selector)
            try:
                count = min(locator.count(), 10)
            except Exception:
                continue
            for i in range(count):
                item = locator.nth(i)
                try:
                    if item.is_visible():
                        return item
                except Exception:
                    continue
        return None

    def login_if_needed(self) -> None:
        self.page.goto(self.base_url + "/", wait_until="domcontentloaded")
        self.page.wait_for_timeout(800)

        password_field = self._first_visible_css(PASSWORD_SELECTORS)
        if password_field is not None:
            username_field = self._first_visible_css(USERNAME_SELECTORS)
            if username_field and password_field:
                login_btn = None
                for text in LOGIN_BUTTON_TEXTS:
                    loc = self.page.get_by_role("button", name=text, exact=False)
                    try:
                        for i in range(min(loc.count(), 5)):
                            cand = loc.nth(i)
                            if cand.is_visible():
                                login_btn = cand
                                break
                    except Exception:
                        continue
                    if login_btn:
                        break
                if not login_btn:
                    login_btn = self._first_visible_css(("button[type=submit]", ".ant-btn-primary", "button"))

                if username_field and login_btn:
                    username_field.fill(self.username)
                    password_field.fill(self.password)
                    login_btn.click()
                    self.page.wait_for_timeout(1200)

        self.storage_state.parent.mkdir(parents=True, exist_ok=True)
        self.context.storage_state(path=str(self.storage_state))

    def derive_module_name(self, route_id: str) -> str:
        if route_id == "login":
            return "Authentication"
        elif route_id == "doctor":
            return "Doctor Portal"
        elif route_id in ("study-list", "userTable"):
            return "Study List"
        elif route_id == "viewer-modal":
            return "Viewer Window"
        elif "userTable" in route_id:
            clean = route_id.replace("userTable/", "").replace("%E8%83%B8%E9%83%A8DR", "Chest DR")
            return f"Study List ({clean})"
        return route_id.replace("_", " ").replace("-", " ").title()

    def extract_and_classify_route(self, route_id: str, module_name: str | None = None) -> None:
        self.routes_visited.add(route_id)
        if not module_name:
            module_name = self.derive_module_name(route_id)

        self.page.wait_for_timeout(1000)
        current_url = self.page.url
        current_title = ""
        try:
            current_title = self.page.title().strip()
        except Exception:
            current_title = ""

        # Audit page title string itself
        if current_title and len(current_title) >= 2:
            title_cls, title_exp, title_note = classify_string(current_title)
            if title_cls != "technical-term":
                title_key = (route_id, "<title>", current_title)
                if title_key not in self.seen_keys:
                    self.seen_keys.add(title_key)
                    title_fid = generate_finding_id(route_id, "<title>", current_title)
                    self.findings.append(Finding(
                        finding_id=title_fid,
                        route=route_id,
                        module_name=module_name,
                        page_url=current_url,
                        page_title=current_title,
                        element_selector="<title>",
                        text_observed=current_title,
                        classification=title_cls,
                        expected_indonesian=title_exp,
                        quality_note=title_note,
                        screenshot_path="",
                        fullpage_screenshot_path=f"{self.screenshots_dir.name}/fullpage_{route_id}.png",
                    ))

        elements_data = self.page.evaluate(
            """() => {
                const items = [];
                const isVisible = (el) => {
                    if (!el) return false;
                    const style = window.getComputedStyle(el);
                    return style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0';
                };
                
                const processElement = (el) => {
                    if (!isVisible(el)) return;
                    
                    // Check if inside patient table data cell
                    const isTd = Boolean(el.closest('td') && !el.closest('th'));
                    
                    let text = '';
                    let sel = el.tagName.toLowerCase();
                    if (el.id) {
                        sel += '#' + el.id;
                    } else if (el.className && typeof el.className === 'string') {
                        const cls = el.className.split(/\\s+/).filter(c => c && !c.includes('active')).slice(0, 2);
                        if (cls.length > 0) sel += '.' + cls.join('.');
                    }

                    if (el.tagName === 'INPUT' || el.tagName === 'TEXTAREA') {
                        const ph = el.getAttribute('placeholder');
                        if (ph) {
                            text = ph.trim();
                            sel += '[placeholder]';
                        }
                    } else if (el.tagName === 'IMG') {
                        const alt = el.getAttribute('alt') || el.getAttribute('title');
                        if (alt) text = alt.trim();
                    } else {
                        // Check direct child text nodes
                        for (let c of el.childNodes) {
                            if (c.nodeType === Node.TEXT_NODE) {
                                const t = c.textContent.trim();
                                if (t.length >= 2) {
                                    text = t;
                                    break;
                                }
                            }
                        }
                        if (!text) {
                            const aria = el.getAttribute('aria-label') || el.getAttribute('title');
                            if (aria) text = aria.trim();
                        }
                    }

                    if (text && text.length >= 2) {
                        items.push({
                            selector: sel,
                            text: text,
                            isTd: isTd
                        });
                    }
                };

                const all = document.querySelectorAll('span, div, button, a, input, th, td, label, p, h1, h2, h3, h4, h5, h6, li, option');
                for (let el of all) {
                    processElement(el);
                }
                return items;
            }"""
        )

        self.screenshots_dir.mkdir(parents=True, exist_ok=True)

        # Capture full-page screenshot for this route
        fullpage_rel_path = ""
        fullpage_filename = f"fullpage_{route_id}.png"
        fullpage_full_path = self.screenshots_dir / fullpage_filename
        try:
            self.page.screenshot(path=str(fullpage_full_path), full_page=True)
            fullpage_rel_path = f"{self.screenshots_dir.name}/{fullpage_filename}"
        except Exception:
            fullpage_rel_path = ""

        for item in elements_data:
            text = item.get("text", "").strip()
            selector = item.get("selector", "")
            is_td = item.get("isTd", False)

            self.strings_inspected += 1

            # Exclude patient data inside td
            if is_td:
                continue

            classification, expected, quality_note = classify_string(text)
            if classification == "technical-term":
                continue

            key = (route_id, selector, text)
            if key in self.seen_keys:
                continue

            self.seen_keys.add(key)
            finding_id = generate_finding_id(route_id, selector, text)

            # Capture element screenshot
            screenshot_rel_path = ""
            screenshot_filename = f"{route_id}_{len(self.findings)+1:03d}_{finding_id[:8]}.png"
            screenshot_full_path = self.screenshots_dir / screenshot_filename

            try:
                # Find matching Playwright locator
                loc = self.page.get_by_text(text, exact=True).first
                if loc and loc.is_visible():
                    loc.screenshot(path=str(screenshot_full_path))
                    screenshot_rel_path = f"{self.screenshots_dir.name}/{screenshot_filename}"
            except Exception:
                screenshot_rel_path = ""

            self.findings.append(Finding(
                finding_id=finding_id,
                route=route_id,
                module_name=module_name,
                page_url=current_url,
                page_title=current_title,
                element_selector=selector,
                text_observed=text,
                classification=classification,
                expected_indonesian=expected,
                quality_note=quality_note,
                screenshot_path=screenshot_rel_path,
                fullpage_screenshot_path=fullpage_rel_path,
            ))

    def discover_and_audit_navigation(self) -> None:
        """Safely discovers and audits all accessible navigation links and sub-tabs."""
        # Detect menu links and sub-tabs
        nav_elements = self.page.evaluate(
            """() => {
                const links = [];
                const selectors = ['.ant-menu-item', '.ant-menu-submenu-title', 'a[href*="#/"]', '.ant-tabs-tab'];
                const unsafeKeywords = ['delete', 'hapus', 'remove', 'submit', 'kirim', 'save', 'simpan', 'edit', 'ubah', 'download', 'unduh', 'export', 'ekspor', 'logout', 'exit', 'out'];
                
                for (const sel of selectors) {
                    for (const el of document.querySelectorAll(sel)) {
                        const txt = (el.textContent || '').trim().toLowerCase();
                        const href = el.getAttribute('href') || '';
                        
                        // Safety check: skip mutating/destructive links
                        if (unsafeKeywords.some(k => txt.includes(k) || href.toLowerCase().includes(k))) {
                            continue;
                        }
                        
                        links.push({
                            text: el.textContent.trim(),
                            href: href,
                            isTab: el.classList.contains('ant-tabs-tab'),
                        });
                    }
                }
                return links;
            }"""
        )

        for item in nav_elements:
            text = item.get("text", "").strip()
            href = item.get("href", "").strip()
            is_tab = item.get("isTab", False)

            if not text or len(text) < 2:
                continue

            route_name = text.lower().replace(" ", "-")
            if is_tab:
                # Safely click tab if visible
                try:
                    tab_loc = self.page.get_by_role("tab", name=text, exact=False).first
                    if tab_loc and tab_loc.is_visible():
                        tab_loc.click(timeout=3000)
                        self.page.wait_for_timeout(800)
                        self.extract_and_classify_route(f"tab-{route_name}", module_name=f"Tab: {text}")
                except Exception:
                    pass
            elif href and "#/" in href:
                # Navigate to hash route
                hash_part = href[href.index("#/"):].strip()
                if hash_part and hash_part not in ("/login", "#/login"):
                    try:
                        self.page.goto(f"{self.base_url}/{hash_part}", wait_until="domcontentloaded")
                        self.page.wait_for_timeout(800)
                        self.extract_and_classify_route(f"nav-{route_name}", module_name=f"Navigation: {text}")
                    except Exception:
                        pass

    def audit_pdf_reports(self) -> None:
        """Audits downloadable AI diagnostic PDF reports by parsing text content using pypdf."""
        reports_dir = Path("reports/downloaded_reports")
        reports_dir.mkdir(parents=True, exist_ok=True)

        try:
            import pypdf
        except ImportError:
            return

        # Scan for PDF links or trigger PDF download if button exists
        pdf_buttons = self.page.locator("button, a").filter(has_text=re.compile(r"(Report|Laporan|PDF)", re.IGNORECASE))
        btn_count = min(pdf_buttons.count(), 3)

        for i in range(btn_count):
            try:
                btn = pdf_buttons.nth(i)
                if not btn.is_visible():
                    continue

                btn_text = btn.text_content() or ""
                # Skip destructive buttons
                if any(k in btn_text.lower() for k in ("delete", "hapus", "edit", "submit")):
                    continue

                with self.page.expect_download(timeout=5000) as download_info:
                    btn.click(timeout=3000)

                download = download_info.value
                suggested_filename = download.suggested_filename or f"report_{i+1}.pdf"
                pdf_full_path = reports_dir / suggested_filename
                download.save_as(str(pdf_full_path))

                # Parse PDF text
                reader = pypdf.PdfReader(str(pdf_full_path))
                for page_idx, page in enumerate(reader.pages, start=1):
                    raw_text = page.extract_text() or ""
                    for line in raw_text.splitlines():
                        line_clean = line.strip()
                        if len(line_clean) < 2:
                            continue

                        classification, expected, quality_note = classify_string(line_clean)
                        if classification == "technical-term":
                            continue

                        route_id = "pdf-report"
                        selector = f"pdf:page_{page_idx}"
                        key = (route_id, selector, line_clean)
                        if key in self.seen_keys:
                            continue

                        self.seen_keys.add(key)
                        finding_id = generate_finding_id(route_id, selector, line_clean)

                        self.findings.append(Finding(
                            finding_id=finding_id,
                            route=route_id,
                            module_name="AI Diagnostic Report (PDF)",
                            page_url=f"file://{pdf_full_path.resolve()}",
                            page_title=f"PDF Report: {suggested_filename}",
                            element_selector=selector,
                            text_observed=line_clean,
                            classification=classification,
                            expected_indonesian=expected,
                            quality_note=quality_note,
                            screenshot_path="",
                            fullpage_screenshot_path="",
                        ))
            except Exception:
                continue

    def run_audit(self) -> None:
        self.viewer_modal_visited = False
        self.login_if_needed()

        # 1. Login route
        self.page.goto(f"{self.base_url}/#/login", wait_until="domcontentloaded")
        self.extract_and_classify_route("login", module_name="Authentication")

        # 2. Doctor landing route
        self.page.goto(f"{self.base_url}/{DEFAULT_DOCTOR_HASH}", wait_until="domcontentloaded")
        self.extract_and_classify_route("doctor", module_name="Doctor Portal")
        self.discover_and_audit_navigation()

        # 3. Study list route
        self.page.goto(f"{self.base_url}/{DEFAULT_LIST_HASH}", wait_until="domcontentloaded")
        self.extract_and_classify_route("study-list", module_name="Study List (Chest DR)")
        self.discover_and_audit_navigation()

        # 4. Standalone DR Viewer page route
        dr_viewer_url = f"{self.base_url}/view/dr/index.html/viewer?action=viewer&type=CR&sid=58&pacs=fei&aiCalcId=50"
        try:
            self.page.goto(dr_viewer_url, wait_until="domcontentloaded")
            self.page.wait_for_timeout(2000)
            self.extract_and_classify_route("dr-viewer-standalone", module_name="Chest DR Intelligent Analysis")
            self.audit_pdf_reports()
        except Exception as exc:
            print(f"[dr-viewer-standalone] Failed to load standalone viewer page: {exc}")

        # 5. Viewer modal route (click first visible study row)
        try:
            self.page.goto(f"{self.base_url}/{DEFAULT_LIST_HASH}", wait_until="domcontentloaded")
            self.page.wait_for_timeout(1000)
            rows = self.page.locator("tbody tr")
            row_count = 0
            try:
                row_count = rows.count()
            except Exception:
                pass

            visible_row = None
            for i in range(min(row_count, 5)):
                try:
                    candidate = rows.nth(i)
                    if candidate.is_visible():
                        visible_row = candidate
                        break
                except Exception:
                    continue

            if visible_row is None:
                print(
                    "[viewer-modal] No visible study row found in study list "
                    "— skipping viewer modal route."
                )
            else:
                try:
                    visible_row.click(timeout=10000)
                    self.page.wait_for_timeout(1500)
                    self.extract_and_classify_route("viewer-modal", module_name="Viewer Window")
                    self.viewer_modal_visited = True
                except Exception as exc:
                    print(
                        f"[viewer-modal] Failed to open viewer modal: "
                        f"{type(exc).__name__}: {exc}"
                    )
        except Exception as outer_exc:
            print(f"[viewer-modal] Unexpected error during row detection: {outer_exc}")


def main() -> int:
    parser = argparse.ArgumentParser(description="AI-PACS Indonesian Localization Audit Tool")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL, help="Base URL of AI-PACS")
    parser.add_argument("--credentials", default="credential.txt", help="Credentials file path")
    parser.add_argument("--storage-state", default="auth-state.json", help="Playwright storage state file")
    parser.add_argument("--output", default="reports/localization_report.xlsx", help="Output Excel report path")
    parser.add_argument("--screenshots-dir", default="reports/screenshots", help="Directory for evidence screenshots")
    parser.add_argument("--headed", action="store_true", help="Run browser in headed mode")
    parser.add_argument("--timeout-ms", type=int, default=30000, help="Default Playwright timeout in ms")
    parser.add_argument("--probe-only", action="store_true", help="Probe PACS reachability only and exit")

    args = parser.parse_args()

    if args.probe_only:
        probe = probe_http_origin(args.base_url)
        if probe["ok"]:
            print(f"PACS probe success: {probe['url']} (status {probe['status']})")
            return 0
        else:
            print(f"PACS probe failed: {probe['error']}", file=sys.stderr)
            return 2

    try:
        username, password = load_credentials(args.credentials)
    except Exception as exc:
        print(f"Credential load error: {exc}", file=sys.stderr)
        return 2

    auditor = LocalizationAuditor(
        base_url=args.base_url,
        username=username,
        password=password,
        storage_state=args.storage_state,
        screenshots_dir=args.screenshots_dir,
        headed=args.headed,
        timeout_ms=args.timeout_ms,
    )

    try:
        auditor.start()
        auditor.run_audit()
    except Exception as exc:
        print(f"Audit run error: {exc}", file=sys.stderr)
        auditor.close()
        return 2
    finally:
        auditor.close()

    # Counts
    counts = {
        "not-indonesian": 0,
        "mixed": 0,
        "quality-issue": 0,
        "uncertain": 0,
    }
    for f in auditor.findings:
        if f.classification in counts:
            counts[f.classification] += 1

    summary = {
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "base_url": args.base_url,
        "routes_visited": len(auditor.routes_visited),
        "strings_inspected": auditor.strings_inspected,
        "findings_total": len(auditor.findings),
        "findings_not_indonesian": counts["not-indonesian"],
        "findings_mixed": counts["mixed"],
        "findings_quality_issue": counts["quality-issue"],
        "findings_uncertain": counts["uncertain"],
        "viewer_modal_visited": getattr(auditor, "viewer_modal_visited", False),
    }

    try:
        export_to_excel(auditor.findings, summary, args.output)
    except Exception as exc:
        print(f"Failed to export report: {exc}", file=sys.stderr)
        return 2

    print("--- Localization Audit Summary ---")
    print(f"Routes Visited    : {summary['routes_visited']}")
    print(f"Strings Inspected : {summary['strings_inspected']}")
    print(f"Findings Total    : {summary['findings_total']}")
    print(f"  - Not Indonesian: {counts['not-indonesian']}")
    print(f"  - Mixed         : {counts['mixed']}")
    print(f"  - Quality Issue : {counts['quality-issue']}")
    print(f"  - Uncertain     : {counts['uncertain']}")
    print(f"Report exported to: {args.output}")

    return 1 if len(auditor.findings) > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
